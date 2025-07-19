from token import STAR
from tokenize import group
from urllib.request import Request
import openpyxl
import re
from datetime import datetime
from collections import defaultdict
import os as os
import platform

class PrinterRequest:
    def __init__(self, id, serial_number, list_count, request_date, end_date):
        self.ID = id
        self.SerialNumber = serial_number
        self.ListCount = list_count
        self.RequestDate = request_date
        self.EndDate = end_date

class NoRequest:
    def __init__(self, id, request_date, FIO):
        self.ID = id
        self.RequestDate = request_date
        self.FIO = FIO
        
countCells = 1
countTR1 = 0
countTR2 = 0
countOT6 = 0
countOT7 = 0
countOT3 = 0
countSCS8 = 0
countSCS9 = 0
countARM11 = 0
countTK13 = 0
countTK14 = 0
countNO16 = 0
frodList6 = []
frodList7 = []
frodList16 = []
startPrinter6 = []
startPrinter7 = []
startNO16 = defaultdict(list)
dublList = []

while True:
    try:
        path = input("Укажите путь к таблице: ").replace('"', '').strip()
        workbook = openpyxl.load_workbook(filename = f"{path}")
        worksheet = workbook.worksheets[0]
        break
    except: 
        print("\nНеверный путь! \nПоддерживаемый формат файла: .xlsx, .xlsm, .xltx, .xltm\n")
        continue

# === ПЕРВЫЙ ФРОД ===
def check_TR1(i):
    global countTR1
    if worksheet.cell(row=i, column=12).value == "Решено полностью":
        cell_value = str(worksheet.cell(row=i, column=10).value)
        if any(marker in cell_value for marker in ["АРМ_ТР_202", "НОУТБУК_ТР_202", "РАЗРАБОТЧИК_ТР_202", "МРМ_ТР_202", "АЗАРМ_ТР_202"]):
            cell_value = str(worksheet.cell(row=i, column=11).value)
            pattern0 = r"###.*.*#1###"
            pattern1 = r"ТМЦ###.*.*#1###"
            pattern2 = r"##SerNo:.*.{,10}\s*\$"

            match0 = re.search(pattern0, cell_value)
            match1 = re.search(pattern1, cell_value)
            match2 = re.search(pattern2, cell_value)
            if not (match0 or match1 or match2 or ("#фрод_тр" in str(worksheet.cell(row=i, column=10).value))):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_1 ТР"
                countTR1 += 1
                return

        if "ОТ_ТР_202" in str(worksheet.cell(row=i, column=10).value):
            cell_value = str(worksheet.cell(row=i, column=11).value)
            pattern0 = r"###.*.*#1###"
            pattern1 = r"ТМЦ###.*.*#1###"
            pattern2 = r"#Nomer .*.*#"

            match0 = re.search(pattern0, cell_value)
            match1 = re.search(pattern1, cell_value)
            match2 = re.search(pattern2, cell_value)
            if not (match0 or match1 or match2):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_1 ТР"
                countTR1 += 1
                return

        if "ТК_ТР_202" in str(worksheet.cell(row=i, column=10).value):
            cell_value = str(worksheet.cell(row=i, column=11).value)
            pattern0 = r"ТМЦ###.*.*#1#"

            match0 = re.search(pattern0, cell_value)
            if not (match0 or ("SB3" in cell_value) or ("SB1" in cell_value)):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_1 ТР"
                countTR1 += 1
                return
        
        if any(marker in cell_value for marker in ["АРМ_ТР_202", "НОУТБУК_ТР_202", "РАЗРАБОТЧИК_ТР_202", "МРМ_ТР_202", "АЗАРМ_ТР_202", "_ТК_202", "ТК_ТР_202"]):
            cell_value = str(worksheet.cell(row=i, column=11).value)
            pattern0 = r"support.sberbank-service.ru/sm/.*?id"

            match0 = re.search(pattern0, cell_value)
            if not (match0):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_1 ТР"
                countTR1 += 1
                return

# === ВТОРОЙ ФРОД ===
def check_TR2(i):
    cell_value = str(worksheet.cell(row=i, column=11).value)

    if any(marker in cell_value for marker in ["АРМ_ТР_202", "НОУТБУК_ТР_202", "РАЗРАБОТЧИК_ТР_202"]):
        if (("#ARM010207" in str(worksheet.cell(row=i, column=11).value)) or "#фрод_тр" in str(worksheet.cell(row=i, column=10).value)):
            pattern0 = r"#Имя нового ...:\s*.{3}\s*-\s*.{3}\s*-\s*.{7}\s*"

            match0 = re.search(pattern0, cell_value)
            if not (match0):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_2 ТР"
                countTR2 += 1

    if any(marker in cell_value for marker in ["_ТК_202", "ТК_ТР_202"]):
        pattern0 = r".{3}\s*-\s*.{3}\s*-\s*.{7}"
        
        match0 = re.search(pattern0, cell_value)
        if not (match0):
            worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_2 ТР"
            countTR2 += 1

# === ТРЕТИЙ ФРОД ===
def check_OT3(i):
    global countOT3
    if "Поле Location на устройстве:" in str(worksheet.cell(row=i, column=10).value):
        if "PREDICT_ОТ_Заполнение поля Location на принтере/МФУ" in str(worksheet.cell(row=i, column=9).value):
            pattern = r"\d{4}/\d{4};\s*\S+,\s*\S+,\s*\S+;\s*Et.\s*\w+;\s*Pm.\s*\w+"
            if "Переданные вложения по задаче можно получить по ссылке:" not in str(worksheet.cell(row=i, column=11).value):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_3 ОТ"
                countOT3 += 1

# === ЧЕТВЁРТЫЙ ФРОД ===
def check_OT4(i):
    cell_value = str(worksheet.cell(row=i, column=10).value)
    pattern0 = r"Счётчик отпечатанных страниц: ([1-9]\d{3,})\$"
    pattern1 = r"Количество лотков для бумаги: [4-9]\d{0,}\$"
    pattern2 = r"Серийный:\s*([^\s$]+)\s*\$"
    
    match0 = re.search(pattern0, cell_value)
    match1 = re.search(pattern1, cell_value)
    match2 = re.search(pattern2, cell_value)   
    if match0 and match1 and match2:
        request_date = worksheet.cell(row=i, column=4).value
        end_date = worksheet.cell(row=i, column=5).value
        if (isinstance(request_date, datetime) and isinstance(end_date, datetime)):
            startPrinter7.append(PrinterRequest(
                str(worksheet.cell(row=i, column=1).value),
                match2.group(1),
                match0.group(1),
                request_date,
                end_date
            ))

# === ШЕСТОЙ ФРОД ===
def check_OT6(i):
    cell_value = str(worksheet.cell(row=i, column=10).value)
    pattern0 = r"На устройстве необходимо заменить расходные материалы: / (.*)\$"
    pattern1 = r"Серийный:\s*([^\s$]+)\s*\$"

    match0 = re.search(pattern0, cell_value)
    match1 = re.search(pattern1, cell_value)
    if match0 and match1:
        request_date = worksheet.cell(row=i, column=4).value
        end_date = worksheet.cell(row=i, column=5).value
        if (isinstance(request_date, datetime) and isinstance(end_date, datetime)):
            startPrinter6.append(PrinterRequest(
                str(worksheet.cell(row=i, column=1).value),
                match1.group(1),
                match0.group(1),
                request_date,
                end_date
            ))

# === СОРТИРОВКА ДЛЯ 6 ФРОДА ===
def check_OT6_final():
    global countOT6, frodList6
    grouped = defaultdict(list)
    for request in startPrinter6:
        grouped[request.SerialNumber].append(request)
    
    for serial, requests in grouped.items():
        sorted_dates = sorted(requests, key=lambda x: x.RequestDate)
        filtered_group = []
        
        if sorted_dates:
            filtered_group.append(sorted_dates[0])
            
            for i in range(1, len(sorted_dates)):
                if (((sorted_dates[i].RequestDate - filtered_group[-1].RequestDate).days >= 4) and (filtered_group[-1].EndDate < sorted_dates[i].RequestDate) and (filtered_group[-1].ListCount == sorted_dates[i].ListCount)):
                    filtered_group.append(sorted_dates[i])
        
        for req in filtered_group:
            frodList6.append(req.ID)

    countOT6 = len(frodList6)

# === СЕДЬМОЙ ФРОД ===
def check_OT7(i):
    cell_value = str(worksheet.cell(row=i, column=10).value)
    pattern0 = r"Счётчик отпечатанных страниц: ([1-9]\d{3,})\$"
    pattern1 = r"Количество лотков для бумаги: [4-9]\d{0,}\$"
    pattern2 = r"Серийный:\s*([^\s$]+)\s*\$"
    
    match0 = re.search(pattern0, cell_value)
    match1 = re.search(pattern1, cell_value)
    match2 = re.search(pattern2, cell_value)   
    if match0 and match1 and match2:
        request_date = worksheet.cell(row=i, column=4).value
        end_date = worksheet.cell(row=i, column=5).value
        if (isinstance(request_date, datetime) and isinstance(end_date, datetime)):
            startPrinter7.append(PrinterRequest(
                str(worksheet.cell(row=i, column=1).value),
                match2.group(1),
                match0.group(1),
                request_date,
                end_date
            ))

# === СОРТИРОВКА ДЛЯ 7 ФРОДА ===
def check_OT7_final():
    global countOT7, frodList7
    grouped = defaultdict(list)
    for request in startPrinter7:
        grouped[request.SerialNumber].append(request)
    
    for serial, requests in grouped.items():
        sorted_dates = sorted(requests, key=lambda x: x.RequestDate)
        filtered_group = []
        
        if sorted_dates:
            filtered_group.append(sorted_dates[0])
            
            for i in range(1, len(sorted_dates)):
                if (((sorted_dates[i].RequestDate - filtered_group[-1].RequestDate).days >= 4) and (filtered_group[-1].EndDate < sorted_dates[i].RequestDate) and ((int(sorted_dates[i].ListCount) - int(filtered_group[-1].ListCount)) < 1000)):
                    filtered_group.append(sorted_dates[i])

        for req in filtered_group:
            frodList7.append(req.ID)

    countOT7 = len(frodList7)

# === ВОСЬМОЙ ФРОД ===
def check_SCS8(i):
    global countSCS8
    if worksheet.cell(row=i, column=12).value == "Решено полностью":
        cell_value = str(worksheet.cell(row=i, column=10).value)
        if "фот" in cell_value and "#FIXSKS" not in cell_value:
            if "Переданные вложения по задаче можно получить по ссылке:" not in str(worksheet.cell(row=i, column=11).value):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_8 СКС"
                countSCS8 += 1

# === ДЕВЯТЫЙ ФРОД ===
def check_SCS9(i):
    global countSCS9
    if "#ПриемкаСКС" in str(worksheet.cell(row=i, column=10).value) and worksheet.cell(row=i, column=2).value == "Выполнен":
        cell_value = str(worksheet.cell(row=i, column=11).value)
        if not any(marker in cell_value for marker in [
            "#ARM12", "#ARM510305", "#ARM510306", "#ARM510307", "#ARM510308"
        ]):
            worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_9 СКС"
            countSCS9 += 1

# === ТРИНАДЦАТЫЙ ФРОД ===
def check_TK13(i):
    global countTK13
    if worksheet.cell(row=i, column=12).value == "Решено полностью":
        if "Переданные вложения по задаче можно получить по ссылке:" not in str(worksheet.cell(row=i, column=11).value):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_13 ТК"
                countTK13 += 1

# === ЧЕТЫРНАДЦАТЫЙ ФРОД ===
def check_TK14(i):
    global countTK14
    if worksheet.cell(row=i, column=12).value != "Решено полностью":
        if "Переданные вложения по задаче можно получить по ссылке:" in str(worksheet.cell(row=i, column=11).value):
                worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_14 ТК"
                countTK14 += 1

# === ШЕСТНАДЦАТЫЙ ФРОД ==
def check_NО16(i):
    global startNO16
    cell_value = str(worksheet.cell(row=i, column=11).value)
    pattern0 = r"Дубль заявки:\s*(.*)\s*\$"
    pattern1 = r"Исполнитель:\s*(.*)\s*"
    
    match0 = re.search(pattern0, cell_value)
    match1 = re.search(pattern1, cell_value)

    if (match0 and match1):
        startNO16[match0.group(1).strip()].append(NoRequest(
            str(worksheet.cell(row=i, column=1).value),
            str(worksheet.cell(row=i, column=4).value),
            match1.group(1)))

# === СОРТИРОВКА ДЛЯ 16 ФРОДА ===
def check_NO16_final():
    global countNO16
    for i in dublList:
        for y in startNO16.get(i.ID):
            if ((i.RequestDate > y.RequestDate) and (i.FIO == y.FIO)):
                frodList16.append(y.ID)
                countNO16 += 1

# === ПОДСЧЁТ КОЛ-ВА ЗАПИСЕЙ ===
while worksheet.cell(row=countCells, column=1).value is not None:
    countCells += 1

worksheet.cell(row=1, column=14).value = "Найденные ФРОДы"

# === ОСНОВНОЙ КОД ПРОГРАММЫ ===
for i in range(2, countCells):
    worksheet.cell(row=i, column=14).value = ""
    cell_value = str(worksheet.cell(row=i, column=7).value)
    
    if "(CI00306594)" in cell_value:
        check_TR1(i)
        check_TR2(i)
    
    if "(CI00306597)" in cell_value:
        check_OT3(i)
        check_OT6(i)
        check_OT7(i)
    
    if "(CI01547267)" in cell_value:
        check_SCS8(i)
        check_SCS9(i)
    
# === ОДИНАДЦАТЫЙ ФРОД ===
    if (("Решено полностью" in str(worksheet.cell(row=i, column=12).value)) and ("ARM12" in str(worksheet.cell(row=i, column=11).value)) and ("ARM120116" not in str(worksheet.cell(row=i, column=11).value))):
        countARM11 += 1
        worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_11 АРМ"
# ========================

    if ("тонкий клиент") in str(worksheet.cell(row=i, column=10).value).lower():
        check_TK13(i)
        check_TK14(i)
    
    if ("Дублирование" in str(worksheet.cell(row=i, column=12).value)):
        check_NО16(i)

check_OT6_final()
check_OT7_final()

for i in range(1, countCells):
    if str(worksheet.cell(row=i, column=1).value) in frodList6:
        worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_6 ОТ"

    if str(worksheet.cell(row=i, column=1).value) in frodList7:
        worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_7 ОТ"

    if (str(worksheet.cell(row=i, column=1).value) in startNO16.keys()):
        pattern0 = r"Исполнитель:\s*(.*)\s*"
        match0 = re.search(pattern0, str(worksheet.cell(row=i, column=11).value))
        
        dublList.append(NoRequest(
            str(worksheet.cell(row=i, column=1).value),
            str(worksheet.cell(row=i, column=4).value),
            match0.group(1)))

check_NO16_final()

for i in range(1, countCells):
    if str(worksheet.cell(row=i, column=1).value) in frodList16:
        worksheet.cell(row=i, column=14).value = worksheet.cell(row=i, column=14).value + "\n" + "ФРОД_16 НО"

workbook.save(path)
print(f"\nНайдено ФРОД_1: {countTR1}")
print(f"Найдено ФРОД_2: {countTR2}")
print(f"Найдено ФРОД_3: {countOT3}")
print(f"Найдено ФРОД_4: Не найдено")
print(f"Найдено ФРОД_5: Не найдено")
print(f"Найдено ФРОД_6: {countOT6}")
print(f"Найдено ФРОД_7: {countOT7}")
print(f"Найдено ФРОД_8: {countSCS8}")
print(f"Найдено ФРОД_9: {countSCS9}")
print(f"Найдено ФРОД_10: Не найдено")
print(f"Найдено ФРОД_11: {countARM11}")
print(f"Найдено ФРОД_12: Не найдено")
print(f"Найдено ФРОД_13: {countTK13}")
print(f"Найдено ФРОД_14: {countTK14}")
print(f"Найдено ФРОД_15: Не найдено")
print(f"Найдено ФРОД_16: {countNO16}")
print(f"Найдено ФРОД_17: Не найдено")
print(f"Найдено ФРОД_18: Не найдено")
print(f"Найдено ФРОД_19: Не найдено")